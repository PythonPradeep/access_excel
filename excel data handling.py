#import needed libraries
import openpyxl
#select the file path
file_path = "/home/pradeepthamizh/Desktop/data handling/task.xlsx"
#use load_workbook for 
file = openpyxl.load_workbook(file_path)
#use active for
new_sheet = file.active

#select the value of the student in the file
#student1
cell_obj = new_sheet['A2': 'F2']
for cell1, cell2, cell3, cell4, cell5, cell6 in cell_obj:
    # print(cell1.value, cell2.value, cell3.value, cell4.value, cell5.value, cell6.value)
     marks = cell2.value + cell3.value + cell4.value + cell5.value + cell6.value

#student2
cell_obj = new_sheet['A3': 'F3']
for cell1, cell2, cell3, cell4, cell5, cell6 in cell_obj:
    # print(cell1.value, cell2.value, cell3.value, cell4.value, cell5.value, cell6.value)
    marks2 = cell2.value + cell3.value + cell4.value + cell5.value + cell6.value

#student3
cell_obj = new_sheet['A4': 'F4']
for cell1, cell2, cell3, cell4, cell5, cell6 in cell_obj:
    # print(cell1.value, cell2.value, cell3.value, cell4.value, cell5.value, cell6.value)
    marks3 = cell2.value + cell3.value + cell4.value + cell5.value + cell6.value

#student4
cell_obj = new_sheet['A5': 'F5']
for cell1, cell2, cell3, cell4, cell5, cell6 in cell_obj:
    # print(cell1.value, cell2.value, cell3.value, cell4.value, cell5.value, cell6.value)
    marks4 = cell2.value + cell3.value + cell4.value + cell5.value + cell6.value

#student5
cell_obj = new_sheet['A6': 'F6']
for cell1, cell2, cell3, cell4, cell5, cell6 in cell_obj:
    # print(cell1.value, cell2.value, cell3.value, cell4.value, cell5.value, cell6.value)
    marks5 = cell2.value + cell3.value + cell4.value + cell5.value + cell6.value


# print(marks, marks2, marks3, marks4, marks5)

new_sheet['G2'] = marks
new_sheet['G3'] = marks2
new_sheet['G4'] = marks3
new_sheet['G5'] = marks4
new_sheet['G6'] = marks5

file.save('task.xlsx')

#test workings...
# row = new_sheet.max_row
# column = new_sheet.max_column

# my_cell_obj = new_sheet.cell(row = 1, column = 1)
# print(my_cell_obj)

# student1 = new_sheet.cell(row = 2, column = 1) 
# print(student1.value)  

#s1_m1 = student1_mark1

# s1_m1 = new_sheet.cell(row = 2, column = 2)
# print(s1_m1.value) 
# s1_m2 = new_sheet.cell(row = 2, column = 3)
# print(s1_m2.value)
# s1_m3 = new_sheet.cell(row = 2, column = 4)
# print(s1_m3.value) 
# s1_m4 = new_sheet.cell(row = 2, column = 5)
# print(s1_m4.value) 
# s1_m5 = new_sheet.cell(row = 2, column = 6)
# print(s1_m5.value) 


# for i in range(1, column + 1): 
#     s1 = new_sheet.cell(row = 2, column = i) 
#     print(s1.value, end = " ")
# print("\n")

# for i in range(1, column + 1): 
#     s1 = new_sheet.cell(row = 3, column = i) 
#     print(s1.value, end = " ")
# print("\n")

# for i in range(1, column + 1): 
#     s1 = new_sheet.cell(row = 4, column = i) 
#     print(s1.value, end = " ")
# print("\n")

# for i in range(1, column + 1): 
#     s1 = new_sheet.cell(row = 5, column = i) 
#     print(s1.value, end = " ")
# print("\n")

# for i in range(1, column + 1): 
#     s1 = new_sheet.cell(row = 6, column = i) 
#     print(s1.value, end = " ")

# total1 = s1_m1.value + s1_m2.value + s1_m3.value +s1_m4.value + s1_m5.value
# print(total1)
